#!/usr/bin/env python3
"""
Parses the "All" tab of the Network-wide Content Scope and Sequences
workbook into data/all_grades.json for the portal - one entry per grade
(TK, KN, 1st-8th), so the Grade Level filter has real options to switch
between.

Usage:
    python3 parse_grades.py path/to/workbook.xlsx ../data/all_grades.json

Why "All" and not a per-grade tab (e.g. "3rd"): those per-grade tabs are
separately-maintained copies that drift out of sync with "All", the tab
every grade's data actually lives in - see README.md ("Why the data comes
from the 'All' tab") for specifics. "All" is the source of truth here.

Row-label convention in column A:
    "<Grade(s)>: <Subject> Section"  -> one row per unit/section (topic, dates, standards)
    "<Grade(s)>: <Subject> Lessons"  -> one row per individual lesson
    "<Grade(s)>: Assessment"         -> one row per assessment window
    "Network-wide Key Dates"         -> shared calendar events (holidays, PD days, etc.)
A row applies to a grade if that grade's token appears in the
comma-separated "Grade Level" column (column B) - not just by parsing the
label prefix, since shared rows (Values Block, Assessment, Key Dates) list
multiple grades there. A single row can and does apply to several grades
at once (e.g. "3, 4, 5: Values Block"); it's copied into each grade's
dataset, tagged with the same original row number (_row) so the portal
can de-duplicate it when multiple grades are selected together.

Two data-quality gaps in "All" needed generalized (not grade-3-specific)
handling - see the functions below for details:
  - classify_label(): several grades (1, 2, 4, 5, 6, 7, 8) have "Section"
    rows missing their " Section" suffix entirely (e.g. "2: Math" instead
    of "2: Math Section"). Distinguished from real Lessons rows by
    whether a lesson_number is present.
  - reconstruct_missing_section_numbers(): "All" has a native Section
    column, and most grades have it filled in - but it's blank for every
    row in grades 3, 6, 7, and 8. For those, unit numbers are
    reconstructed from row order (see the function docstring).
"""
import sys
import json
import re
import datetime
from pathlib import Path
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    sys.exit("Run: pip install openpyxl --break-system-packages")

SOURCE_SHEET = "All"

# Ordered so the Grade Level filter (and this file's JSON key order) reads
# TK, KN, 1st..8th instead of whatever order dict insertion happens to hit.
GRADE_TOKENS = ["TK", "KN", "1", "2", "3", "4", "5", "6", "7", "8"]
GRADE_KEY = {
    "TK": "TK", "KN": "KN", "1": "1st", "2": "2nd", "3": "3rd", "4": "4th",
    "5": "5th", "6": "6th", "7": "7th", "8": "8th",
}

KNOWN_SUBJECTS = {"Math", "ELA", "SLA", "Science", "Social Studies"}

COLUMNS = [
    None, "combined_grade_subject", "grade_level", "subject_area",
    "language_of_instruction", "curriculum", "section_number", "topic",
    "lesson_number", "lesson_objective", "language_objective", "resources",
    "start_date", "end_date", "focus_mastery_standards", "spiral_standards",
    "assessments",
]

LABEL_RE = re.compile(r"^(.*?):\s*(.+)$")
MODULE_HEADER_RE = re.compile(r"^Module\s+\d+:")
BARE_DIGIT_RE = re.compile(r"^\d+$")
DATE_RE = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})$")


def parse_date_to_iso(s):
    """Normalizes an M/D/YYYY text date (as opposed to a real Excel date
    cell, which cell_to_value already turns into ISO via .isoformat()) to
    the same "YYYY-MM-DDT00:00:00" shape the portal's toDate() expects.
    Without this, a raw "8/17/2026" string reaches the browser as-is and
    "8/17/2026T00:00:00" is not a valid JS Date, silently dropping that
    lesson/section off the Timeline and Weekly Planning entirely."""
    m = DATE_RE.match(s) if s else None
    if not m:
        return None
    month, day, year = (int(x) for x in m.groups())
    return f"{year:04d}-{month:02d}-{day:02d}T00:00:00"
LESSON_TOKEN_RE = re.compile(r"^L?\d+$", re.IGNORECASE)


def looks_like_lesson_number(value):
    """Real lesson_number values in this sheet are short tokens like "L1",
    "L12/13" (combined lessons), or "Review" - never a topic sentence.
    Used to sanity-check the lesson_number peek in main(): some bare-label
    Section rows put the topic/module name in this exact column slot (see
    fix_bare_section_topic_slot), and a plain length cutoff isn't reliable
    - some real topics are short ("Habitats") and some are borderline-long
    - but the *shape* of a real lesson_number is distinctive and doesn't
    overlap with topic text at all."""
    if value is None:
        return False
    s = str(value).strip()
    if not s:
        return False
    if s.lower() in ("review", "flex day", "flex days"):
        return True
    parts = re.split(r"\s*[/|]\s*", s)
    return all(LESSON_TOKEN_RE.match(p.strip()) for p in parts if p.strip())


def fix_mojibake(s):
    """A chunk of the Social Studies rows (around the Studies Weekly import)
    have Spanish text that was double-encoded at some point before it ever
    reached this sheet: UTF-8 bytes got misread as MacRoman, turning "ó"
    into "√≥", "¿" into "¬ø", etc. This reverses that specific corruption
    where it's detectable, and leaves everything else untouched."""
    if not s or ("√" not in s and "¬" not in s):
        return s
    try:
        fixed = s.encode("mac_roman").decode("utf-8")
        if "√" not in fixed and "¬" not in fixed:
            return fixed
    except (UnicodeEncodeError, UnicodeDecodeError):
        pass
    return s


def cell_to_value(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    if v is None:
        return None
    return fix_mojibake(str(v).strip())


def detect_column_shift(ws, r):
    """A subset of rows (found in Grade 6-8 Math and Science so far - e.g.
    "Eureka Math Squared" landing in column D) skip the Language of
    Instruction column (D) entirely during entry, which shifts every field
    after it one column left of where the standard schema expects it.
    Detected generically: column D should only ever hold blank/English/
    Spanish - anything else means this row is shifted."""
    v = ws.cell(row=r, column=4).value
    return 0 if v in (None, "English", "Spanish") else 1


def row_dict(ws, r, ncols=16):
    shift = detect_column_shift(ws, r)
    d = {"_row": r}
    for c in range(1, ncols + 1):
        if shift and c == 4:
            # Language of Instruction was skipped entirely on this row -
            # that's *why* everything after it is shifted - so there's no
            # sheet cell to read it from.
            d[COLUMNS[c]] = None
            continue
        src_col = c - shift if c >= 5 else c
        cell = ws.cell(row=r, column=src_col)
        d[COLUMNS[c]] = cell_to_value(cell.value)
        # The "Resources" cell is often a hyperlinked doc title/name (e.g.
        # "Copy of Integrity 10/13" linking to a Google Doc) rather than a
        # visible URL - openpyxl exposes that as cell.hyperlink, separate
        # from the cell's displayed text/value.
        if COLUMNS[c] == "resources" and cell.hyperlink is not None:
            d["resources_url"] = cell.hyperlink.target
    fix_resources_slot_date_shift(d, ws, r, shift)
    return d


def fix_resources_slot_date_shift(d, ws, r, shift):
    """227 rows (nearly all of Grade 4's ELA/SLA lessons, plus a handful of
    Grade 1-2 Science sections) genuinely have two different dates - but
    they were entered one column early: the real start_date landed in the
    Resources column (K) and the real end_date landed in what the standard
    schema calls start_date (L), leaving the schema's end_date column (M)
    blank. Confirmed narrow and safe to detect purely by type: a raw
    (pre-string-conversion) datetime in the Resources cell is never a
    legitimate resource - real resource values are always hyperlinked doc
    titles (plain text). Every row matching this pattern also has a real
    date in the following column and a blank one after that, so the
    remap is unambiguous."""
    resources_col = 11 - shift if 11 >= 5 else 11
    raw = ws.cell(row=r, column=resources_col).value
    if not isinstance(raw, (datetime.datetime, datetime.date)):
        return
    # The real start_date is this Resources-slot value. What row_dict
    # already read into "start_date" (from the standard schema's
    # start_date column, one column over) is actually the real end_date -
    # re-read that same physical column explicitly rather than trusting
    # the schema label on it.
    end_col = 12 - shift if 12 >= 5 else 12
    d["start_date"] = cell_to_value(raw)
    d["end_date"] = cell_to_value(ws.cell(row=r, column=end_col).value)
    d["resources"] = None
    d.pop("resources_url", None)


def classify_label(label, lesson_number):
    """Returns (kind, subject) where kind is one of KeyDate/Assessment/
    Section/Lessons, or (None, None) if the label doesn't match anything
    recognized (e.g. "#REF!" or a stray row)."""
    if not label:
        return None, None
    if label.strip() == "Network-wide Key Dates":
        return "KeyDate", None
    m = LABEL_RE.match(label)
    if not m:
        return None, None
    rest = m.group(2).strip()
    if rest == "Assessment":
        return "Assessment", None
    if rest.endswith(" Lessons"):
        return "Lessons", rest[: -len(" Lessons")].strip()
    if rest.endswith(" Lesson"):
        return "Lessons", rest[: -len(" Lesson")].strip()
    if rest.endswith(" Section"):
        return "Section", rest[: -len(" Section")].strip()
    if rest.endswith(" Block"):
        return "Section", rest
    # Several grades (1, 2, 4, 5, 6, 7, 8) have Section rows with the
    # suffix dropped entirely - just "2: Math" instead of "2: Math
    # Section". A lesson_number distinguishes an actual lesson row that
    # happens to share the same bare naming from a genuine section/unit
    # placeholder (verified against the sheet: bare "Math"/"SLA"/"Science"
    # rows split roughly 20-80% lesson-vs-section by this signal).
    if rest in KNOWN_SUBJECTS:
        return ("Lessons" if lesson_number else "Section"), rest
    return None, None


def fix_mislabeled_section_header(kind, d):
    """One specific mislabel found in Grade 3's Math sequence: Module 2 is
    labeled "... Lessons" (proper suffix) but is actually a section header
    - no lesson_number, no start_date, topic reads "Module 2: ...". Only
    reclassify on that exact combination so it doesn't misfire on
    legitimate dateless/numberless lesson placeholders elsewhere in the
    sheet (e.g. "Flex Days" rows, which don't match the "Module N:" topic
    pattern)."""
    if kind != "Lessons":
        return kind
    if d.get("lesson_number") or d.get("start_date"):
        return kind
    topic = d.get("topic") or ""
    if MODULE_HEADER_RE.match(topic):
        return "Section"
    return kind


def fix_compact_curriculum_slot(kind, d):
    """Grade 6-8 Science, and some Grade 6-8 SLA rows, enter the section
    number one column early - in the Curriculum slot - leaving the real
    Section Number column blank (verified: "6: Science" / "6: SLA" rows
    like curriculum='1', section_number=None). Detected purely by content,
    not by subject or by detect_column_shift's language-column check
    (which doesn't fire here - column D is legitimately blank on these
    rows): a bare-digit "curriculum" with no section_number is never a
    real curriculum name, so it must be a misplaced section number."""
    if kind != "Section":
        return d
    curriculum = d.get("curriculum")
    if curriculum and BARE_DIGIT_RE.match(curriculum) and not d.get("section_number"):
        d["section_number"] = curriculum
        d["curriculum"] = None
    return d


def fix_compact_topic_dates_standards(kind, d, ws, r):
    """A related gap in the same Grade 6-8 Science/SLA rows: when topic,
    dates, and standards ARE filled in (e.g. some Grade 6 SLA rows), they
    were packed into the columns immediately after the section number
    (topic in G, dates in H/I, standards in J) instead of their normal
    positions (topic in G, dates in L/M, standards in N) - skipping the
    four lesson-only columns (lesson_number/lesson_objective/
    language_objective/resources) that these lesson-less Section rows
    never use. Only fires when the standard-position read left topic/
    dates/standards all blank AND the compact columns actually hold
    something, so it can never overwrite a normally-populated row (this
    is why Grade 6 Math's identically-shaped but genuinely-empty rows are
    unaffected)."""
    if kind != "Section":
        return d
    if d.get("topic") or d.get("start_date") or d.get("end_date") or d.get("focus_mastery_standards"):
        return d
    raw_topic = cell_to_value(ws.cell(row=r, column=7).value)
    raw_c1 = cell_to_value(ws.cell(row=r, column=8).value)
    raw_c2 = cell_to_value(ws.cell(row=r, column=9).value)
    raw_c3 = cell_to_value(ws.cell(row=r, column=10).value)
    if not (raw_topic or raw_c1 or raw_c2 or raw_c3):
        return d
    d["topic"] = raw_topic
    d["start_date"] = parse_date_to_iso(raw_c1)
    d["end_date"] = parse_date_to_iso(raw_c2)
    d["focus_mastery_standards"] = raw_c3
    return d


def fix_bare_section_topic_slot(kind, d):
    """Bare-label Section rows ("2: Math", not "3: Math Section") use yet a
    third column convention: section_number is correctly in column F, but
    the topic/module name lands one column late, in column H (the
    lesson_number slot), leaving column G (the real topic column) blank.
    Verified across Grade 2 Math/Science/SLA. Safe because a genuine
    Section row never has a real lesson_number to begin with - only
    reassigns when topic is blank and something is sitting in the
    lesson_number slot; a normally-populated topic (e.g. the suffixed
    "3: Math Section" rows, which use column G correctly) is untouched."""
    if kind != "Section":
        return d
    if d.get("topic") is None and d.get("lesson_number") is not None:
        d["topic"] = d["lesson_number"]
    # A genuine Section row never has a real lesson_number - clear it here
    # (rather than only in the branch above) so a row where topic was
    # already correctly populated elsewhere (e.g. by
    # fix_compact_topic_dates_standards) doesn't end up with a leftover
    # duplicate of that same value sitting in lesson_number.
    d["lesson_number"] = None
    return d


def applicable_grades(grade_level_value):
    if grade_level_value is None:
        return []
    tokens = {t.strip() for t in str(grade_level_value).split(",")}
    return [g for g in GRADE_TOKENS if g in tokens]


def reconstruct_missing_section_numbers(grade_data):
    """Most grades have "All" tab's native Section column (F) filled in -
    trust it as-is when present. Grades 3, 6, 7, and 8 have it blank for
    every row, though, so for those, unit numbers are reconstructed from
    row order instead: within each (subject, curriculum) pair, a Section
    row starts a new unit, and every Lessons row after it (until the next
    Section row for that pair) inherits that number. Mutates sections/
    lessons in place; called once per grade after that grade's rows have
    all been collected in original sheet order."""
    sections, lessons = grade_data["sections_by_subject"], grade_data["lessons_by_subject"]
    counters, current = {}, {}
    all_rows = []
    for subj, arr in sections.items():
        if subj == "Values Block":
            continue
        all_rows += [("Section", subj, d) for d in arr]
    for subj, arr in lessons.items():
        all_rows += [("Lessons", subj, d) for d in arr]
    all_rows.sort(key=lambda t: t[2]["_row"])

    for kind, subj, d in all_rows:
        if d.get("section_number"):
            key = (subj, d.get("curriculum"))
            current[key] = d["section_number"]
            continue
        key = (subj, d.get("curriculum"))
        if kind == "Section":
            counters[key] = counters.get(key, 0) + 1
            d["section_number"] = str(counters[key])
            current[key] = d["section_number"]
        else:
            d["section_number"] = current.get(key)


def main():
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    src, out_path = sys.argv[1], Path(sys.argv[2])
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb[SOURCE_SHEET]

    by_grade = OrderedDict(
        (GRADE_KEY[g], {"sections_by_subject": {}, "lessons_by_subject": {}, "assessments": [], "key_dates": []})
        for g in GRADE_TOKENS
    )

    for r in range(2, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        # Peek lesson_number (column 8, or 7 if this row is column-shifted -
        # see detect_column_shift) before building the full dict, since
        # classify_label needs it to disambiguate bare "N: Subject" labels.
        # Sanity-check the shape (see looks_like_lesson_number): several
        # bare-label Section rows (Grade 2 Science/Math/SLA, Grade 6-8
        # Science/SLA) put the topic/module name in this exact shift-
        # adjusted slot instead of a real lesson_number, which would
        # otherwise falsely look truthy and misclassify a Section row as
        # Lessons.
        shift = detect_column_shift(ws, r)
        lesson_number = ws.cell(row=r, column=8 - shift).value
        if not looks_like_lesson_number(lesson_number):
            lesson_number = None
        kind, subject = classify_label(label, lesson_number)
        if kind is None:
            continue
        d = row_dict(ws, r)
        kind = fix_mislabeled_section_header(kind, d)
        d = fix_compact_curriculum_slot(kind, d)
        d = fix_compact_topic_dates_standards(kind, d, ws, r)
        d = fix_bare_section_topic_slot(kind, d)
        grades = applicable_grades(d.get("grade_level"))
        if not grades:
            continue

        for g in grades:
            bucket = by_grade[GRADE_KEY[g]]
            row_copy = dict(d)
            if kind == "KeyDate":
                bucket["key_dates"].append(row_copy)
            elif kind == "Assessment":
                bucket["assessments"].append(row_copy)
            elif kind == "Section":
                bucket["sections_by_subject"].setdefault(subject, []).append(row_copy)
            elif kind == "Lessons":
                bucket["lessons_by_subject"].setdefault(subject, []).append(row_copy)

    for grade_data in by_grade.values():
        reconstruct_missing_section_numbers(grade_data)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(by_grade, indent=2, ensure_ascii=False))

    for g, grade_data in by_grade.items():
        n_lessons = sum(len(v) for v in grade_data["lessons_by_subject"].values())
        n_sections = sum(len(v) for v in grade_data["sections_by_subject"].values())
        print(f"{g}: lessons={n_lessons}, sections={n_sections}, "
              f"assessments={len(grade_data['assessments'])}, key_dates={len(grade_data['key_dates'])}")
    print("wrote", out_path)


if __name__ == "__main__":
    main()
